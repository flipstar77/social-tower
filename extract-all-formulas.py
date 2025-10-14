"""
Comprehensive formula extraction from TheTowerofTobi Excel file
This will extract ALL formulas from key sheets to understand the calculation logic
"""
import openpyxl
import json
import sys

def extract_sheet_formulas(sheet, sheet_name, key_ranges=None):
    """
    Extract formulas from a sheet
    If key_ranges provided, focus on those areas
    Otherwise scan entire used range
    """
    print(f"\n{'='*80}")
    print(f"EXTRACTING FORMULAS FROM: {sheet_name}")
    print(f"{'='*80}")

    formulas = {}
    values = {}

    if key_ranges:
        # Scan specific ranges
        for range_name, cell_range in key_ranges.items():
            print(f"\nScanning {range_name}: {cell_range}")
            for row in sheet[cell_range]:
                for cell in row:
                    if cell.value:
                        cell_ref = f"{cell.column_letter}{cell.row}"
                        if str(cell.value).startswith('='):
                            formulas[cell_ref] = str(cell.value)
                            print(f"  Formula: {cell_ref} = {str(cell.value)[:80]}...")
                        else:
                            values[cell_ref] = cell.value
    else:
        # Scan entire used range
        print(f"Scanning entire sheet (may take a while)...")
        max_row = min(sheet.max_row, 200)  # Limit to first 200 rows
        max_col = min(sheet.max_column, 200)  # Limit to first 200 columns

        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell_ref = f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"
                    if str(cell.value).startswith('='):
                        formulas[cell_ref] = str(cell.value)
                    else:
                        values[cell_ref] = cell.value

            # Progress indicator
            if row_idx % 20 == 0:
                print(f"  Scanned {row_idx}/{max_row} rows... ({len(formulas)} formulas found)")

    print(f"\n✅ Found {len(formulas)} formulas and {len(values)} values")

    return formulas, values

# Load workbook
print("Loading Excel workbook...")
print("(This may take a minute for large files)")
wb = openpyxl.load_workbook(
    'server/database/migrations/TheTowerofTobi.xlsx',
    data_only=False,
    read_only=True
)

print(f"\n✅ Loaded workbook with {len(wb.sheetnames)} sheets")
print(f"Sheets: {', '.join(wb.sheetnames)}")

# Extract from key sheets
sheets_to_extract = {
    'eDamage': {
        'inputs': 'C3:D20',           # Stat labels and values
        'lab_effects': 'K3:K20',      # Lab multipliers
        'card_mastery': 'BL5:BR5',    # Card mastery levels
        'constants': 'BY5:CK5',       # Constants
        'calculation': 'EP5:ES5',     # Final eDamage calculation
        'components': 'BU5:CG5',      # Calculation components
    },
    'eHP': {
        'inputs': 'C3:D15',
        'calculation': 'P5:S5',
    },
    'eEcon': {
        'inputs': 'C3:D15',
        'calculation': 'P5:S5',
    },
    'Lab Researches': None,  # Will scan entire sheet for lab data
}

all_extractions = {}

for sheet_name, ranges in sheets_to_extract.items():
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        formulas, values = extract_sheet_formulas(sheet, sheet_name, ranges)
        all_extractions[sheet_name] = {
            'formulas': formulas,
            'values': values,
            'formula_count': len(formulas),
            'value_count': len(values)
        }
    else:
        print(f"\n⚠️  Sheet '{sheet_name}' not found!")

# Save to JSON files
print(f"\n{'='*80}")
print("SAVING RESULTS")
print(f"{'='*80}")

for sheet_name, data in all_extractions.items():
    filename = f"formulas-{sheet_name.replace(' ', '_')}.json"
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"✅ Saved {sheet_name} to {filename}")

# Create summary
summary = {
    'total_sheets': len(all_extractions),
    'sheets': {
        name: {
            'formulas': data['formula_count'],
            'values': data['value_count']
        }
        for name, data in all_extractions.items()
    }
}

with open('formula-extraction-summary.json', 'w') as f:
    json.dump(summary, f, indent=2)

print(f"\n✅ Summary saved to formula-extraction-summary.json")

print(f"\n{'='*80}")
print("EXTRACTION COMPLETE!")
print(f"{'='*80}")
print(f"\nTotal formulas extracted: {sum(d['formula_count'] for d in all_extractions.values())}")
print(f"Total values extracted: {sum(d['value_count'] for d in all_extractions.values())}")

"""
Quick check of Excel file structure
"""
import openpyxl

print("Loading workbook...")
wb = openpyxl.load_workbook('server/database/migrations/TheTowerofTobi.xlsx', data_only=False, read_only=True)

print(f"\nSheets: {wb.sheetnames}\n")

# Check eDamage sheet
if 'eDamage' in wb.sheetnames:
    sheet = wb['eDamage']

    print("Checking key cells in eDamage sheet:")
    print("-" * 60)

    # Check a few specific cells that should have formulas
    cells_to_check = [
        ('C5', 'Stat label row 5'),
        ('D5', 'Value row 5'),
        ('EP5', 'Possible calc column'),
        ('ER5', 'Possible calc column'),
        ('ES5', 'Possible calc column'),
    ]

    for cell_ref, description in cells_to_check:
        cell = sheet[cell_ref]
        value = cell.value if cell.value else "[EMPTY]"
        is_formula = str(value).startswith('=') if value and value != "[EMPTY]" else False

        print(f"\n{cell_ref} ({description}):")
        print(f"  Value: {str(value)[:100]}")
        if is_formula:
            print(f"  [FORMULA DETECTED]")

print("\nDone!")

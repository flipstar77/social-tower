"""
Extract formulas from TheTowerofTobi Excel file
"""
import openpyxl
import json

# Load the workbook
wb = openpyxl.load_workbook('server/database/migrations/TheTowerofTobi.xlsx', data_only=False)

print("=" * 80)
print("EXCEL WORKBOOK ANALYSIS")
print("=" * 80)

# List all sheets
print(f"\nAvailable sheets: {wb.sheetnames}")

# Focus on eDamage sheet first
if 'eDamage' in wb.sheetnames:
    sheet = wb['eDamage']
    print("\n" + "=" * 80)
    print("ANALYZING eDamage SHEET")
    print("=" * 80)

    # Find cells with formulas in the calculation area
    print("\n📊 Looking for formulas in calculation columns (E5:DZ10)...\n")

    formulas_found = {}

    for row_idx in range(3, 10):  # Rows 3-9
        for col_idx in range(40, 130):  # Columns 40-130 (where calculations are)
            cell = sheet.cell(row=row_idx, column=col_idx)

            if cell.value and str(cell.value).startswith('='):
                cell_ref = f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"
                formulas_found[cell_ref] = cell.value
                print(f"  {cell_ref}: {cell.value[:100]}...")  # First 100 chars

    print(f"\n✅ Found {len(formulas_found)} formulas")

    # Look specifically at the eDamage output cell area
    print("\n" + "=" * 80)
    print("SEARCHING FOR eDamage OUTPUT (around column 127, row 5)")
    print("=" * 80)

    for row_idx in range(4, 6):
        for col_idx in range(125, 135):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell_ref = f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"

            if cell.value:
                if str(cell.value).startswith('='):
                    print(f"\n🎯 {cell_ref} (FORMULA):")
                    print(f"   Formula: {cell.value}")
                else:
                    print(f"\n📈 {cell_ref} (VALUE): {cell.value}")

    # Look at the stat input cells
    print("\n" + "=" * 80)
    print("STAT INPUT CELLS (around rows 3-20, columns 50-70)")
    print("=" * 80)

    for row_idx in range(3, 20):
        # Column C (index 3) has labels
        label_cell = sheet.cell(row=row_idx, column=3)
        # Column D (index 4) has values
        value_cell = sheet.cell(row=row_idx, column=4)

        if label_cell.value:
            value = value_cell.value if value_cell.value else "N/A"
            is_formula = str(value).startswith('=') if value else False

            print(f"Row {row_idx}: {label_cell.value} = {value if not is_formula else '[FORMULA]'}")
            if is_formula:
                print(f"         Formula: {value[:150]}...")

    # Save all formulas to JSON
    output = {
        "sheet": "eDamage",
        "formulas": formulas_found,
        "total_formulas": len(formulas_found)
    }

    with open('edmg-formulas.json', 'w') as f:
        json.dump(output, f, indent=2)

    print(f"\n✅ Formulas saved to edmg-formulas.json")

else:
    print("\n❌ eDamage sheet not found!")
    print(f"Available sheets: {wb.sheetnames}")

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
